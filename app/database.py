import sqlite3
import os
import shutil
import csv
from datetime import date, datetime, time
from typing import List, Optional, Tuple, Dict, Any
from contextlib import contextmanager

from .models import (
    Patient, Visit, Photo, Note, DATA_DIR,
    PHOTO_STATUS_PENDING, PHOTO_STATUS_TAKEN, PHOTO_STATUS_MISSING, PHOTO_STATUS_SKIPPED
)

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False


@contextmanager
def get_db():
    db_path = os.path.join(DATA_DIR, 'clinic.db')
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        conn.close()


class Database:
    DB_VERSION = 2

    @staticmethod
    def init():
        with get_db() as conn:
            c = conn.cursor()

            c.execute('''
                CREATE TABLE IF NOT EXISTS patients (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    phone TEXT,
                    medical_record_number TEXT,
                    age INTEGER,
                    gender TEXT,
                    treatment_plan TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')

            c.execute('''
                CREATE TABLE IF NOT EXISTS visits (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    patient_id INTEGER NOT NULL,
                    visit_date DATE NOT NULL,
                    appointment_time TIME,
                    stage_code TEXT DEFAULT 'month1',
                    notes TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (patient_id) REFERENCES patients(id) ON DELETE CASCADE
                )
            ''')

            c.execute('''
                CREATE TABLE IF NOT EXISTS photos (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    visit_id INTEGER NOT NULL,
                    position_code TEXT NOT NULL,
                    file_path TEXT NOT NULL,
                    file_name TEXT NOT NULL,
                    status TEXT DEFAULT 'pending',
                    taken_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    imported INTEGER DEFAULT 0,
                    FOREIGN KEY (visit_id) REFERENCES visits(id) ON DELETE CASCADE
                )
            ''')

            c.execute('''
                CREATE TABLE IF NOT EXISTS notes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    photo_id INTEGER NOT NULL,
                    content TEXT NOT NULL,
                    x REAL DEFAULT 0,
                    y REAL DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (photo_id) REFERENCES photos(id) ON DELETE CASCADE
                )
            ''')

            c.execute('''
                CREATE TABLE IF NOT EXISTS db_version (
                    version INTEGER PRIMARY KEY
                )
            ''')

            c.execute('SELECT version FROM db_version')
            row = c.fetchone()
            if not row:
                c.execute('INSERT INTO db_version (version) VALUES (?)', (Database.DB_VERSION,))
            else:
                current_version = row['version']
                if current_version < Database.DB_VERSION:
                    Database._migrate(c, current_version, Database.DB_VERSION)

    @staticmethod
    def _migrate(c: sqlite3.Cursor, from_version: int, to_version: int):
        if from_version < 2:
            try:
                c.execute("ALTER TABLE patients ADD COLUMN medical_record_number TEXT")
            except sqlite3.OperationalError:
                pass
            try:
                c.execute("ALTER TABLE visits ADD COLUMN appointment_time TIME")
            except sqlite3.OperationalError:
                pass
            try:
                c.execute("ALTER TABLE photos ADD COLUMN status TEXT DEFAULT 'pending'")
            except sqlite3.OperationalError:
                pass

            c.execute("UPDATE photos SET status = ? WHERE status IS NULL", (PHOTO_STATUS_TAKEN,))
            c.execute("UPDATE db_version SET version = ?", (to_version,))

    @staticmethod
    def add_patient(patient: Patient) -> int:
        with get_db() as conn:
            c = conn.cursor()
            c.execute('''
                INSERT INTO patients (name, phone, medical_record_number, age, gender, treatment_plan)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (patient.name, patient.phone, patient.medical_record_number,
                  patient.age, patient.gender, patient.treatment_plan))
            return c.lastrowid

    @staticmethod
    def update_patient(patient: Patient) -> bool:
        with get_db() as conn:
            c = conn.cursor()
            c.execute('''
                UPDATE patients SET name=?, phone=?, medical_record_number=?, age=?, gender=?, treatment_plan=?
                WHERE id=?
            ''', (patient.name, patient.phone, patient.medical_record_number,
                  patient.age, patient.gender, patient.treatment_plan, patient.id))
            return c.rowcount > 0

    @staticmethod
    def get_patient(patient_id: int) -> Optional[Patient]:
        with get_db() as conn:
            c = conn.cursor()
            c.execute('SELECT * FROM patients WHERE id = ?', (patient_id,))
            row = c.fetchone()
            if row:
                return Database._row_to_patient(row)
        return None

    @staticmethod
    def find_patient_by_identifier(identifier: str) -> Optional[Patient]:
        with get_db() as conn:
            c = conn.cursor()
            c.execute('''
                SELECT * FROM patients
                WHERE phone = ? OR medical_record_number = ?
                LIMIT 1
            ''', (identifier, identifier))
            row = c.fetchone()
            if row:
                return Database._row_to_patient(row)
        return None

    @staticmethod
    def find_patients_by_name(name: str) -> List[Patient]:
        patients = []
        with get_db() as conn:
            c = conn.cursor()
            c.execute('SELECT * FROM patients WHERE name LIKE ? ORDER BY created_at DESC',
                      (f'%{name}%',))
            for row in c.fetchall():
                patients.append(Database._row_to_patient(row))
        return patients

    @staticmethod
    def get_all_patients() -> List[Patient]:
        patients = []
        with get_db() as conn:
            c = conn.cursor()
            c.execute('SELECT * FROM patients ORDER BY created_at DESC')
            for row in c.fetchall():
                patients.append(Database._row_to_patient(row))
        return patients

    @staticmethod
    def get_today_visits(visit_date: Optional[date] = None) -> List[Tuple[Patient, Visit]]:
        if visit_date is None:
            visit_date = date.today()
        results = []
        with get_db() as conn:
            c = conn.cursor()
            c.execute('''
                SELECT p.id as patient_id, p.name, p.phone, p.medical_record_number,
                       p.age, p.gender, p.treatment_plan, p.created_at as patient_created,
                       v.id as visit_id, v.visit_date, v.appointment_time, v.stage_code,
                       v.notes, v.created_at as visit_created
                FROM patients p
                JOIN visits v ON p.id = v.patient_id
                WHERE v.visit_date = ?
                ORDER BY v.appointment_time IS NULL, v.appointment_time ASC, v.created_at DESC
            ''', (visit_date.isoformat(),))
            for row in c.fetchall():
                patient = Patient(
                    id=row['patient_id'],
                    name=row['name'],
                    phone=row['phone'] if row['phone'] else '',
                    medical_record_number=row['medical_record_number'] if row['medical_record_number'] else '',
                    age=row['age'],
                    gender=row['gender'] if row['gender'] else '',
                    treatment_plan=row['treatment_plan'] if row['treatment_plan'] else '',
                    created_at=datetime.fromisoformat(row['patient_created'])
                )
                visit = Visit(
                    id=row['visit_id'],
                    patient_id=row['patient_id'],
                    visit_date=date.fromisoformat(row['visit_date']),
                    stage_code=row['stage_code'] if row['stage_code'] else 'month1',
                    notes=row['notes'] if row['notes'] else '',
                    created_at=datetime.fromisoformat(row['visit_created'])
                )
                appt_time = row['appointment_time']
                if appt_time:
                    if isinstance(appt_time, str):
                        for fmt in ['%H:%M:%S', '%H:%M']:
                            try:
                                visit.appointment_time = datetime.strptime(appt_time, fmt).time()
                                break
                            except ValueError:
                                continue
                    else:
                        visit.appointment_time = appt_time
                results.append((patient, visit))
        return results

    @staticmethod
    def import_appointments_from_csv(file_path: str, visit_date: Optional[date] = None) -> Tuple[int, List[str]]:
        if visit_date is None:
            visit_date = date.today()

        imported_count = 0
        errors = []

        required_columns = ['姓名']
        optional_columns = ['电话', '手机号', '手机', '病历号', '时间段', '预约时间', '时间', '治疗方案', '疗程', '年龄', '性别']

        try:
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)

                if reader.fieldnames is None:
                    errors.append('CSV文件格式错误，无法读取表头')
                    return 0, errors

                found_cols = [col for col in required_columns if col in reader.fieldnames]
                if not found_cols:
                    errors.append(f'缺少必要列：{required_columns}，当前表头：{reader.fieldnames}')
                    return 0, errors

                for row_num, row in enumerate(reader, start=2):
                    try:
                        name = row.get('姓名', '').strip()
                        if not name:
                            errors.append(f'第{row_num}行：姓名为空，已跳过')
                            continue

                        phone = row.get('电话', row.get('手机号', row.get('手机', ''))).strip()
                        medical_record_number = row.get('病历号', '').strip()
                        treatment_plan = row.get('治疗方案', row.get('疗程', '')).strip()

                        age_str = row.get('年龄', '').strip()
                        age = int(age_str) if age_str.isdigit() else None
                        gender = row.get('性别', '').strip()

                        time_str = row.get('时间段', row.get('预约时间', row.get('时间', ''))).strip()
                        appointment_time = None
                        if time_str:
                            for fmt in ['%H:%M', '%H.%M', '%H点%M分', '%H:%M:%S']:
                                try:
                                    appointment_time = datetime.strptime(time_str, fmt).time()
                                    break
                                except ValueError:
                                    continue

                        patient = None
                        if medical_record_number:
                            patient = Database.find_patient_by_identifier(medical_record_number)
                        if not patient and phone:
                            patient = Database.find_patient_by_identifier(phone)
                        if not patient:
                            existing = Database.find_patients_by_name(name)
                            if len(existing) == 1:
                                patient = existing[0]
                            elif len(existing) > 1:
                                errors.append(f'第{row_num}行：存在多名叫「{name}」的患者，请在CSV中填写电话或病历号区分')
                                continue

                        if not patient:
                            patient = Patient(
                                name=name,
                                phone=phone,
                                medical_record_number=medical_record_number,
                                age=age,
                                gender=gender,
                                treatment_plan=treatment_plan
                            )
                            patient_id = Database.add_patient(patient)
                            patient.id = patient_id
                        else:
                            updated = False
                            if phone and not patient.phone:
                                patient.phone = phone
                                updated = True
                            if medical_record_number and not patient.medical_record_number:
                                patient.medical_record_number = medical_record_number
                                updated = True
                            if age and not patient.age:
                                patient.age = age
                                updated = True
                            if updated:
                                Database.update_patient(patient)

                        existing_visits = Database.get_patient_visits(patient.id)
                        has_today_visit = any(v.visit_date == visit_date for v in existing_visits)

                        if not has_today_visit:
                            visit = Visit(
                                patient_id=patient.id,
                                visit_date=visit_date,
                                appointment_time=appointment_time,
                                stage_code='month1'
                            )
                            Database.add_visit(visit)
                            imported_count += 1
                        else:
                            errors.append(f'第{row_num}行：{name} 今日已有预约，已跳过')

                    except Exception as e:
                        errors.append(f'第{row_num}行处理出错：{str(e)}')
                        continue

        except Exception as e:
            errors.append(f'读取CSV文件失败：{str(e)}')
            return 0, errors

        return imported_count, errors

    @staticmethod
    def import_appointments_from_excel(file_path: str, visit_date: Optional[date] = None) -> Tuple[int, List[str]]:
        if not HAS_OPENPYXL:
            return 0, ['未安装 openpyxl 库，无法读取 Excel 文件。请运行：pip install openpyxl']

        if visit_date is None:
            visit_date = date.today()

        imported_count = 0
        errors = []

        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                errors.append('Excel文件为空')
                return 0, errors

            headers = [str(cell).strip() if cell is not None else '' for cell in rows[0]]

            if '姓名' not in headers:
                errors.append(f'缺少必要列：姓名，当前表头：{headers}')
                return 0, errors

            for row_num, row in enumerate(rows[1:], start=2):
                try:
                    row_dict = {headers[i]: str(row[i]).strip() if row[i] is not None else '' for i in range(len(headers))}

                    name = row_dict.get('姓名', '').strip()
                    if not name:
                        errors.append(f'第{row_num}行：姓名为空，已跳过')
                        continue

                    phone = row_dict.get('电话', row_dict.get('手机号', row_dict.get('手机', ''))).strip()
                    medical_record_number = row_dict.get('病历号', '').strip()
                    treatment_plan = row_dict.get('治疗方案', row_dict.get('疗程', '')).strip()

                    age_str = row_dict.get('年龄', '').strip()
                    age = int(age_str) if age_str.isdigit() else None
                    gender = row_dict.get('性别', '').strip()

                    time_str = row_dict.get('时间段', row_dict.get('预约时间', row_dict.get('时间', ''))).strip()
                    appointment_time = None
                    if time_str:
                        for fmt in ['%H:%M', '%H.%M', '%H点%M分', '%H:%M:%S']:
                            try:
                                appointment_time = datetime.strptime(time_str, fmt).time()
                                break
                            except ValueError:
                                continue
                        if appointment_time is None:
                            try:
                                t = datetime.strptime(time_str, '%H:%M:%S').time()
                                appointment_time = t
                            except ValueError:
                                pass

                    patient = None
                    if medical_record_number:
                        patient = Database.find_patient_by_identifier(medical_record_number)
                    if not patient and phone:
                        patient = Database.find_patient_by_identifier(phone)
                    if not patient:
                        existing = Database.find_patients_by_name(name)
                        if len(existing) == 1:
                            patient = existing[0]
                        elif len(existing) > 1:
                            errors.append(f'第{row_num}行：存在多名叫「{name}」的患者，请填写电话或病历号区分')
                            continue

                    if not patient:
                        patient = Patient(
                            name=name,
                            phone=phone,
                            medical_record_number=medical_record_number,
                            age=age,
                            gender=gender,
                            treatment_plan=treatment_plan
                        )
                        patient_id = Database.add_patient(patient)
                        patient.id = patient_id
                    else:
                        updated = False
                        if phone and not patient.phone:
                            patient.phone = phone
                            updated = True
                        if medical_record_number and not patient.medical_record_number:
                            patient.medical_record_number = medical_record_number
                            updated = True
                        if age and not patient.age:
                            patient.age = age
                            updated = True
                        if updated:
                            Database.update_patient(patient)

                    existing_visits = Database.get_patient_visits(patient.id)
                    has_today_visit = any(v.visit_date == visit_date for v in existing_visits)

                    if not has_today_visit:
                        visit = Visit(
                            patient_id=patient.id,
                            visit_date=visit_date,
                            appointment_time=appointment_time,
                            stage_code='month1'
                        )
                        Database.add_visit(visit)
                        imported_count += 1
                    else:
                        errors.append(f'第{row_num}行：{name} 今日已有预约，已跳过')

                except Exception as e:
                    errors.append(f'第{row_num}行处理出错：{str(e)}')
                    continue

        except Exception as e:
            errors.append(f'读取Excel文件失败：{str(e)}')
            return 0, errors

        return imported_count, errors

    @staticmethod
    def import_appointments_from_xls(file_path: str, visit_date: Optional[date] = None) -> Tuple[int, List[str]]:
        if not HAS_XLRD:
            return 0, ['未安装 xlrd 库，无法读取 XLS 文件。请运行：pip install xlrd==1.2.0']

        if visit_date is None:
            visit_date = date.today()

        imported_count = 0
        errors = []

        try:
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_index(0)

            if sheet.nrows < 1:
                errors.append('XLS文件为空')
                return 0, errors

            headers = [str(sheet.cell_value(0, col)).strip() for col in range(sheet.ncols)]

            if '姓名' not in headers:
                errors.append(f'缺少必要列：姓名，当前表头：{headers}')
                return 0, errors

            for row_num in range(1, sheet.nrows):
                try:
                    row_dict = {}
                    for col in range(sheet.ncols):
                        cell_val = sheet.cell_value(row_num, col)
                        if isinstance(cell_val, float) and cell_val == int(cell_val):
                            cell_val = str(int(cell_val))
                        else:
                            cell_val = str(cell_val).strip()
                        row_dict[headers[col]] = cell_val

                    name = row_dict.get('姓名', '').strip()
                    if not name:
                        errors.append(f'第{row_num + 1}行：姓名为空，已跳过')
                        continue

                    phone = row_dict.get('电话', row_dict.get('手机号', row_dict.get('手机', ''))).strip()
                    medical_record_number = row_dict.get('病历号', '').strip()
                    treatment_plan = row_dict.get('治疗方案', row_dict.get('疗程', '')).strip()

                    age_str = row_dict.get('年龄', '').strip()
                    age = int(float(age_str)) if age_str and age_str.replace('.', '').isdigit() else None
                    gender = row_dict.get('性别', '').strip()

                    time_str = row_dict.get('时间段', row_dict.get('预约时间', row_dict.get('时间', ''))).strip()
                    appointment_time = None
                    if time_str:
                        for fmt in ['%H:%M', '%H.%M', '%H点%M分', '%H:%M:%S']:
                            try:
                                appointment_time = datetime.strptime(time_str, fmt).time()
                                break
                            except ValueError:
                                continue
                        if appointment_time is None:
                            try:
                                if ':' in time_str:
                                    parts = time_str.split(':')
                                    h = int(parts[0])
                                    m = int(parts[1]) if len(parts) > 1 else 0
                                    appointment_time = time(h, m)
                            except (ValueError, IndexError):
                                pass

                    patient = None
                    if medical_record_number:
                        patient = Database.find_patient_by_identifier(medical_record_number)
                    if not patient and phone:
                        patient = Database.find_patient_by_identifier(phone)
                    if not patient:
                        existing = Database.find_patients_by_name(name)
                        if len(existing) == 1:
                            patient = existing[0]
                        elif len(existing) > 1:
                            errors.append(f'第{row_num + 1}行：存在多名叫「{name}」的患者，请填写电话或病历号区分')
                            continue

                    if not patient:
                        patient = Patient(
                            name=name,
                            phone=phone,
                            medical_record_number=medical_record_number,
                            age=age,
                            gender=gender,
                            treatment_plan=treatment_plan
                        )
                        patient_id = Database.add_patient(patient)
                        patient.id = patient_id
                    else:
                        updated = False
                        if phone and not patient.phone:
                            patient.phone = phone
                            updated = True
                        if medical_record_number and not patient.medical_record_number:
                            patient.medical_record_number = medical_record_number
                            updated = True
                        if age and not patient.age:
                            patient.age = age
                            updated = True
                        if updated:
                            Database.update_patient(patient)

                    existing_visits = Database.get_patient_visits(patient.id)
                    has_today_visit = any(v.visit_date == visit_date for v in existing_visits)

                    if not has_today_visit:
                        visit = Visit(
                            patient_id=patient.id,
                            visit_date=visit_date,
                            appointment_time=appointment_time,
                            stage_code='month1'
                        )
                        Database.add_visit(visit)
                        imported_count += 1
                    else:
                        errors.append(f'第{row_num + 1}行：{name} 今日已有预约，已跳过')

                except Exception as e:
                    errors.append(f'第{row_num + 1}行处理出错：{str(e)}')
                    continue

        except Exception as e:
            errors.append(f'读取XLS文件失败：{str(e)}')
            return 0, errors

        return imported_count, errors

    @staticmethod
    def import_appointments(file_path: str, visit_date: Optional[date] = None) -> Tuple[int, List[str]]:
        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.csv':
            return Database.import_appointments_from_csv(file_path, visit_date)
        elif ext == '.xlsx':
            return Database.import_appointments_from_excel(file_path, visit_date)
        elif ext == '.xls':
            return Database.import_appointments_from_xls(file_path, visit_date)
        else:
            return 0, [f'不支持的文件格式：{ext}，请使用 CSV 或 Excel 文件']

    @staticmethod
    def add_visit(visit: Visit) -> int:
        with get_db() as conn:
            c = conn.cursor()
            appt_time = visit.appointment_time.strftime('%H:%M') if visit.appointment_time else None
            c.execute('''
                INSERT INTO visits (patient_id, visit_date, appointment_time, stage_code, notes)
                VALUES (?, ?, ?, ?, ?)
            ''', (visit.patient_id, visit.visit_date.isoformat(), appt_time,
                  visit.stage_code, visit.notes))
            return c.lastrowid

    @staticmethod
    def update_visit(visit: Visit) -> bool:
        with get_db() as conn:
            c = conn.cursor()
            appt_time = visit.appointment_time.strftime('%H:%M') if visit.appointment_time else None
            c.execute('''
                UPDATE visits SET patient_id=?, visit_date=?, appointment_time=?, stage_code=?, notes=?
                WHERE id=?
            ''', (visit.patient_id, visit.visit_date.isoformat(), appt_time,
                  visit.stage_code, visit.notes, visit.id))
            return c.rowcount > 0

    @staticmethod
    def get_patient_visits(patient_id: int) -> List[Visit]:
        visits = []
        with get_db() as conn:
            c = conn.cursor()
            c.execute('''
                SELECT * FROM visits WHERE patient_id = ? ORDER BY visit_date DESC
            ''', (patient_id,))
            for row in c.fetchall():
                visits.append(Database._row_to_visit(row))
        return visits

    @staticmethod
    def get_visit(visit_id: int) -> Optional[Visit]:
        with get_db() as conn:
            c = conn.cursor()
            c.execute('SELECT * FROM visits WHERE id = ?', (visit_id,))
            row = c.fetchone()
            if row:
                return Database._row_to_visit(row)
        return None

    @staticmethod
    def add_photo(photo: Photo) -> int:
        with get_db() as conn:
            c = conn.cursor()
            c.execute('''
                INSERT INTO photos (visit_id, position_code, file_path, file_name, status, taken_at, imported)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                photo.visit_id,
                photo.position_code,
                photo.file_path,
                photo.file_name,
                photo.status,
                photo.taken_at.isoformat(),
                1 if photo.imported else 0
            ))
            return c.lastrowid

    @staticmethod
    def update_photo(photo: Photo) -> bool:
        with get_db() as conn:
            c = conn.cursor()
            c.execute('''
                UPDATE photos SET visit_id=?, position_code=?, file_path=?, file_name=?,
                       status=?, taken_at=?, imported=?
                WHERE id=?
            ''', (
                photo.visit_id, photo.position_code, photo.file_path, photo.file_name,
                photo.status, photo.taken_at.isoformat(), 1 if photo.imported else 0,
                photo.id
            ))
            return c.rowcount > 0

    @staticmethod
    def update_photo_status(photo_id: int, status: str) -> bool:
        with get_db() as conn:
            c = conn.cursor()
            c.execute('UPDATE photos SET status = ? WHERE id = ?', (status, photo_id))
            return c.rowcount > 0

    @staticmethod
    def get_visit_photos(visit_id: int) -> List[Photo]:
        photos = []
        with get_db() as conn:
            c = conn.cursor()
            c.execute('''
                SELECT * FROM photos WHERE visit_id = ? ORDER BY id
            ''', (visit_id,))
            for row in c.fetchall():
                photos.append(Database._row_to_photo(row))
        return photos

    @staticmethod
    def get_visit_photos_by_position(visit_id: int) -> Dict[str, Photo]:
        photos = Database.get_visit_photos(visit_id)
        return {p.position_code: p for p in photos}

    @staticmethod
    def get_photo(photo_id: int) -> Optional[Photo]:
        with get_db() as conn:
            c = conn.cursor()
            c.execute('SELECT * FROM photos WHERE id = ?', (photo_id,))
            row = c.fetchone()
            if row:
                return Database._row_to_photo(row)
        return None

    @staticmethod
    def get_photo_by_position(visit_id: int, position_code: str) -> Optional[Photo]:
        with get_db() as conn:
            c = conn.cursor()
            c.execute('''
                SELECT * FROM photos WHERE visit_id = ? AND position_code = ?
                ORDER BY id DESC LIMIT 1
            ''', (visit_id, position_code))
            row = c.fetchone()
            if row:
                return Database._row_to_photo(row)
        return None

    @staticmethod
    def delete_photo(photo_id: int, delete_file: bool = True) -> bool:
        with get_db() as conn:
            c = conn.cursor()
            photo = Database.get_photo(photo_id)
            if photo and delete_file and os.path.exists(photo.file_path):
                try:
                    os.remove(photo.file_path)
                except Exception as e:
                    print(f"Delete photo file failed: {e}")
            c.execute('DELETE FROM notes WHERE photo_id = ?', (photo_id,))
            c.execute('DELETE FROM photos WHERE id = ?', (photo_id,))
            return c.rowcount > 0

    @staticmethod
    def delete_photo_keep_record(photo_id: int) -> bool:
        with get_db() as conn:
            c = conn.cursor()
            c.execute('''
                UPDATE photos SET status = ?, file_path = '', file_name = ''
                WHERE id = ?
            ''', (PHOTO_STATUS_PENDING, photo_id))
            return c.rowcount > 0

    @staticmethod
    def update_photo_position(photo_id: int, position_code: str) -> bool:
        with get_db() as conn:
            c = conn.cursor()
            c.execute('UPDATE photos SET position_code = ? WHERE id = ?',
                      (position_code, photo_id))
            return c.rowcount > 0

    @staticmethod
    def add_note(note: Note) -> int:
        with get_db() as conn:
            c = conn.cursor()
            c.execute('''
                INSERT INTO notes (photo_id, content, x, y)
                VALUES (?, ?, ?, ?)
            ''', (note.photo_id, note.content, note.x, note.y))
            return c.lastrowid

    @staticmethod
    def get_photo_notes(photo_id: int) -> List[Note]:
        notes = []
        with get_db() as conn:
            c = conn.cursor()
            c.execute('SELECT * FROM notes WHERE photo_id = ? ORDER BY created_at',
                      (photo_id,))
            for row in c.fetchall():
                notes.append(Database._row_to_note(row))
        return notes

    @staticmethod
    def delete_note(note_id: int) -> bool:
        with get_db() as conn:
            c = conn.cursor()
            c.execute('DELETE FROM notes WHERE id = ?', (note_id,))
            return c.rowcount > 0

    @staticmethod
    def copy_photo_to_visit(src_path: str, visit: Visit, patient: Patient, position_code: str) -> Optional[str]:
        visit_dir = visit.get_visit_dir(patient.name, patient.id)
        ext = os.path.splitext(src_path)[1]
        timestamp = datetime.now().strftime('%H%M%S_%f')
        new_filename = f"{position_code}_{timestamp}{ext}"
        dest_path = os.path.join(visit_dir, new_filename)
        try:
            shutil.copy2(src_path, dest_path)
            return dest_path
        except Exception as e:
            print(f"Copy photo failed: {e}")
            return None

    @staticmethod
    def _row_to_patient(row: sqlite3.Row, prefix: str = '') -> Patient:
        p = Patient(
            id=row[f'{prefix}id'] if prefix else row['id'],
            name=row[f'{prefix}name'] if prefix else row['name'],
            phone=row[f'{prefix}phone'] if prefix else row['phone'],
            medical_record_number=row[f'{prefix}medical_record_number'] if prefix else row['medical_record_number'],
            age=row[f'{prefix}age'] if prefix else row['age'],
            gender=row[f'{prefix}gender'] if prefix else row['gender'],
            treatment_plan=row[f'{prefix}treatment_plan'] if prefix else row['treatment_plan'],
            created_at=datetime.fromisoformat(row[f'{prefix}created_at'] if prefix else row['created_at'])
        )
        if prefix == 'patient_id':
            p.id = row['patient_id']
        return p

    @staticmethod
    def _row_to_visit(row: sqlite3.Row, prefix: str = '') -> Visit:
        v = Visit(
            id=row[f'{prefix}id'] if prefix else row['id'],
            patient_id=row[f'{prefix}patient_id'] if prefix else row['patient_id'],
            visit_date=date.fromisoformat(row[f'{prefix}visit_date'] if prefix else row['visit_date']),
            stage_code=row[f'{prefix}stage_code'] if prefix else row['stage_code'],
            notes=row[f'{prefix}notes'] if prefix else row['notes'],
            created_at=datetime.fromisoformat(row[f'{prefix}created_at'] if prefix else row['created_at'])
        )
        appt_time = row[f'{prefix}appointment_time'] if prefix else row['appointment_time']
        if appt_time:
            if isinstance(appt_time, str):
                for fmt in ['%H:%M:%S', '%H:%M']:
                    try:
                        v.appointment_time = datetime.strptime(appt_time, fmt).time()
                        break
                    except ValueError:
                        continue
            else:
                v.appointment_time = appt_time
        if prefix == 'visit_id':
            v.id = row['visit_id']
        return v

    @staticmethod
    def _row_to_photo(row: sqlite3.Row) -> Photo:
        status = row['status'] if 'status' in row.keys() else PHOTO_STATUS_TAKEN
        return Photo(
            id=row['id'],
            visit_id=row['visit_id'],
            position_code=row['position_code'],
            file_path=row['file_path'],
            file_name=row['file_name'],
            status=status,
            taken_at=datetime.fromisoformat(row['taken_at']),
            imported=bool(row['imported'])
        )

    @staticmethod
    def _row_to_note(row: sqlite3.Row) -> Note:
        return Note(
            id=row['id'],
            photo_id=row['photo_id'],
            content=row['content'],
            x=row['x'],
            y=row['y'],
            created_at=datetime.fromisoformat(row['created_at'])
        )
